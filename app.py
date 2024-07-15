from flask import Flask, request, redirect, render_template, send_file, flash, jsonify, url_for
from flask import Flask, send_file
import os
import logging
from database import connect_db
import waitress
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Border, Side
from openpyxl.chart import PieChart, Reference, series
from datetime import datetime
from openpyxl.chart.label import DataLabel, DataLabelList



app = Flask(__name__, template_folder='Templates')
app.secret_key = 'supersecretkey'

UPLOAD_FOLDER = 'uploads'
RESULT_FOLDER = 'results'
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['RESULT_FOLDER'] = RESULT_FOLDER

if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)
if not os.path.exists(RESULT_FOLDER):
    os.makedirs(RESULT_FOLDER)


@app.route('/add_member', methods=['GET', 'POST'])
def add_member():
    if request.method == 'POST':
        try:
            member_id = int(request.form.get('member_id'))
            name = request.form.get('name')
            co_leader = request.form.get('co_leader')

            with connect_db() as conn:  # Context manager for connection
                cursor = conn.cursor()
                cursor.execute("""
                    INSERT INTO members (member_id, name, co_leader) 
                    VALUES (?, ?, ?)
                """, (member_id, name, co_leader))
                conn.commit()  # Commit changes immediately after execution

            flash('Member added successfully!', 'success')
        except ValueError:
            flash('Member ID must be an integer.', 'danger')
        except Exception as e:
            flash(f'Error adding member: {str(e)}', 'danger')

        return redirect(url_for('add_member'))

    return render_template('add_member.html')

@app.route('/reset_results', methods=['POST'])
def reset_results():
    try:
        conn = connect_db()
        cursor = conn.cursor()
        cursor.execute("DELETE FROM results")
        conn.commit()
        conn.close()
        flash('Results have been reset successfully!', 'success')
    except Exception as e:
        logging.error(f"Error resetting results: {e}")
        flash('An error occurred while resetting the results.', 'danger')

    return redirect(url_for('upload_form'))

@app.route('/members')
def view_members():
    try:
        conn = connect_db()
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM members")
        members = cursor.fetchall()
        conn.close()
        return render_template('members.html', members=members)
    except Exception as e:
        logging.error(f"Error fetching members: {e}")
        flash('An error occurred while fetching members.', 'danger')
        return redirect('/')


@app.route('/edit_member/<int:member_id>', methods=['GET', 'POST'])
def edit_member(member_id):
    if request.method == 'POST':
        # Update member details in the  database
        new_name = request.form.get('name')
        new_co_leader = request.form.get('co_leader')
        try:
            conn = connect_db()
            cursor = conn.cursor()
            cursor.execute("""  
                UPDATE members
                SET name=?, co_leader=?
                WHERE member_id=?
            """, (new_name, new_co_leader, member_id))
            conn.commit()
            conn.close()
            flash('Member updated successfully!', 'success')
            return redirect('/members')
        except Exception as e:
            logging.error(f"Error updating member: {e}")
            flash('An error occurred while updating the member.', 'danger')
            return redirect('/members')
    else:
        try:
            conn = connect_db()
            cursor = conn.cursor()
            cursor.execute("SELECT * FROM members WHERE member_id=?", (member_id,))
            member = cursor.fetchone()
            conn.close()
            return render_template('edit_member.html', member=member)
        except Exception as e:
            logging.error(f"Error fetching member details: {e}")
            flash('An error occurred while fetching member details.', 'danger')
            return redirect('/members')

@app.route('/delete_member/<int:member_id>', methods=['GET', 'POST'])
def delete_member(member_id):
    try:
        conn = connect_db()
        cursor = conn.cursor()
        cursor.execute("DELETE FROM members WHERE member_id=?", (member_id,))
        conn.commit()
        conn.close()
        flash('Member deleted successfully!', 'success')
    except Exception as e:
        logging.error(f"Error deleting member: {e}")
        flash('An error occurred while deleting the member.', 'danger')
    return redirect('/members')

@app.route('/upload', methods=['POST'])
def upload_file():
    try:
        if 'report' not in request.files:
            flash('No file part')
            return redirect(request.url)

        report_file = request.files['report']

        if report_file.filename == '':
            flash('No selected file')
            return redirect(request.url)

        report_path = os.path.join(app.config['UPLOAD_FOLDER'], 'report.xlsx')

        report_file.save(report_path)

        result_path = os.path.join(app.config['RESULT_FOLDER'], 'result.xlsx')  # Change file extension to .xlsx

        compare_reports(report_path, result_path)

        return send_file(result_path, as_attachment=True)

    except Exception as e:
        logging.error(f"Error uploading files: {e}")
        flash('An error occurred while processing the files.')
        return redirect(request.url)
def compare_reports(report_path, output_path):
    try:
        conn = connect_db()
        cursor = conn.cursor()

        # Load report file (Excel) with openpyxl
        wb = load_workbook(filename=report_path, read_only=True, data_only=True)
        sheet = wb.active

        submitted_member_ids = []
        for row in sheet.iter_rows(values_only=True):
            try:
                if row and row[0]:  # Assuming member ID is in the first column
                    submitted_member_ids.append(str(row[0]))  # Convert to string if necessary
            except Exception as e:
                logging.warning(f"Error processing row: {e}. Skipping row.")
                continue

        results = []

        # Fetch all members with name and co_leader separately
        cursor.execute("SELECT member_id, name, co_leader FROM members")
        members = cursor.fetchall()

        for member in members:
            member_id, name, co_leader = member
            if str(member_id) in submitted_member_ids:
                status = "Submitted"
            else:
                status = "Not Submitted"
            # Append member details including name to results
            results.append((member_id, name, co_leader, status))

        # Create new workbook and add results to it
        wb_result = Workbook()
        ws_result = wb_result.active
        ws_result.title = "Report Comparison"

        # Set headers
        headers = ['Member ID', 'Name', 'Co-Leader', 'Status']
        ws_result.append(headers)

        # Define styles
        bold_font = Font(bold=True)
        border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        submitted_font = Font(color='006400')  # Dark green
        not_submitted_font = Font(color='FF0000')  # Red

        # Apply header styles
        for cell in ws_result[1]:
            cell.font = bold_font
            cell.border = border

        # Add data and apply styles
        for row_num, (member_id, name, co_leader, status) in enumerate(results, start=2):
            ws_result.append([member_id, name, co_leader, status])
            for col_num in range(1, 5):
                cell = ws_result.cell(row=row_num, column=col_num)
                cell.border = border
                if col_num == 4:  # Status column
                    if status == "Submitted":
                        cell.font = submitted_font
                    else:
                        cell.font = not_submitted_font

        # Save the workbook
        wb_result.save(output_path)

        # Store results in the database
        for member_id, name, co_leader, status in results:
            cursor.execute("""
                INSERT OR REPLACE INTO results (member_id, name, co_leader, result, submission_date)
                VALUES (?, ?, ?, ?, ?)
            """, (member_id, name, co_leader, status, datetime.now()))
        conn.commit()

        conn.close()

    except Exception as e:
        logging.error(f"Error comparing reports: {e}")
        raise

def generate_summary_report(output_path):
    try:
        conn = connect_db()
        cursor = conn.cursor()

        # Fetch the count of submitted and not submitted
        cursor.execute("SELECT result, COUNT(*) FROM results GROUP BY result")
        counts = dict(cursor.fetchall())

        submitted_count = counts.get("Submitted", 0)
        not_submitted_count = counts.get("Not Submitted", 0)

        # Create a new workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Summary Report"

        # Add data
        ws.append(["Status", "Count"])
        ws.append(["Submitted", submitted_count])
        ws.append(["Not Submitted", not_submitted_count])

        # Apply borders and bold the first column
        thin_border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))

        for row in ws.iter_rows(min_row=1, max_row=3, min_col=1, max_col=2):
            for cell in row:
                cell.border = thin_border
                if cell.col_idx == 1:
                    cell.font = Font(bold=True)

        # Create a pie chart
        pie = PieChart()
        labels = Reference(ws, min_col=1, min_row=2, max_row=3)
        data = Reference(ws, min_col=2, min_row=1, max_row=3)
        pie.add_data(data, titles_from_data=True)
        pie.set_categories(labels)
        pie.title = "Submission Status"

        ws.add_chart(pie, "E5")

        # Save the workbook
        wb.save(output_path)

        conn.close()
    except Exception as e:
        logging.error(f"Error generating summary report: {e}")
        raise

def generate_member_summary_report(output_path):
    try:
        conn = connect_db()
        cursor = conn.cursor()

        # Fetch member-wise submission counts
        cursor.execute("""
            SELECT submission_date, name, result
            FROM results
            GROUP BY submission_date, name, result
        """)
        results = cursor.fetchall()

        # Create a new workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Memberwise Summary Report"

        # Add header row
        header = ["Submission Date", "Name", "Status"]

        ws.append(header)

        # Apply bold font to the header row
        for cell in ws[1]:
            cell.font = Font(bold=True)

        # Add data
        for row in results:
            ws.append(row)

        # Apply borders to the entire table and bold the first column
        thin_border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))

        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=3):
            for cell in row:
                cell.border = thin_border
            row[0].font = Font(bold=True)  # Bold the first column

        # Save the workbook
        wb.save(output_path)

        conn.close()
    except Exception as e:
        logging.error(f"Error generating member summary report: {e}")
        raise

@app.route('/')
def upload_form():
    try:
        return render_template('upload.html')
    except Exception as e:
        logging.error(f"Error rendering upload form: {e}")
        return "An error occurred while rendering the form.", 500


@app.route('/download_datewise_summary')
def download_datewise_summary():
    try:
        datewise_summary_path = os.path.join(app.config['RESULT_FOLDER'], 'Datewise_Total_Count.xlsx')
        generate_datewise_summary_report(datewise_summary_path)
        return send_file(datewise_summary_path, as_attachment=True)
    except Exception as e:
        logging.error(f"Error generating datewise summary report: {e}")
        flash('An error occurred while generating the datewise summary report.')
        return redirect('/')

def generate_datewise_summary_report(output_path):
    try:
        conn = connect_db()
        cursor = conn.cursor()

        # Fetch date-wise submission counts
        cursor.execute("""
            SELECT DATE(submission_date) AS submission_date,
                   SUM(CASE WHEN result = 'Submitted' THEN 1 ELSE 0 END) AS submitted_count,
                   SUM(CASE WHEN result = 'Not Submitted' THEN 1 ELSE 0 END) AS not_submitted_count
            FROM results
            GROUP BY DATE(submission_date)
        """)
        results = cursor.fetchall()

        # Create a new workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Datewise Summary Report"

        # Add data with subtotal rows for "Submitted" and "Not Submitted"
        ws.append(["Submission Date", "Submitted Count", "Not Submitted Count"])
        submitted_total = 0
        not_submitted_total = 0

        for row in results:
            submission_date, submitted_count, not_submitted_count = row
            ws.append([submission_date, submitted_count, not_submitted_count])
            submitted_total += submitted_count
            not_submitted_total += not_submitted_count

        # Add subtotal rows
        subtotal_row_idx = ws.max_row + 1
        ws.append(["Subtotal", submitted_total, not_submitted_total])

        # Apply borders and bold formatting
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=3):
            for cell in row:
                cell.border = Border(
                    left=Side(style='thin', color='000000'),
                    right=Side(style='thin', color='000000'),
                    top=Side(style='thin', color='000000'),
                    bottom=Side(style='thin', color='000000')
                )
                if cell.row == 1 or cell.row == subtotal_row_idx:
                    cell.font = Font(bold=True)

        # Insert a pie chart using the subtotal data only
        pie = PieChart()
        labels = Reference(ws, min_col=2, min_row=subtotal_row_idx, max_row=subtotal_row_idx)
        data = Reference(ws, min_col=2, min_row=subtotal_row_idx, max_col=3)
        pie.add_data(data, titles_from_data=True)
        pie.set_categories(labels)
        pie.title = "Submission Status"

        # Show percentages in the pie chart
        pie.dataLabels = DataLabelList()
        pie.dataLabels.showPercent = True

        ws.add_chart(pie, "E1")

        # Save the workbook
        wb.save(output_path)

        conn.close()
    except Exception as e:
        logging.error(f"Error generating datewise summary report: {e}")
        raise
@app.route('/download/<filename>')
def download_file(filename):
    try:
        file_path = os.path.join(app.config['RESULT_FOLDER'], filename)
        return send_file(file_path, as_attachment=True)

    except Exception as e:
        logging.error(f"Error downloading file: {e}")
        flash('An error occurred while downloading the file.')
        return redirect(request.url)

@app.route('/download_member_summary')
def download_member_summary():
    try:
        member_summary_path = os.path.join(app.config['RESULT_FOLDER'], 'Datewise_Member_Report.xlsx')
        generate_member_summary_report(member_summary_path)
        return send_file(member_summary_path, as_attachment=True)
    except Exception as e:
        logging.error(f"Error generating member summary report: {e}")
        flash('An error occurred while generating the member summary report.')
        return redirect('/')


def generate_memberwise_total_report(output_path):
    try:
        conn = connect_db()
        cursor = conn.cursor()

        # Fetch member-wise submission counts
        cursor.execute("""
            SELECT name,
                   SUM(CASE WHEN result = 'Submitted' THEN 1 ELSE 0 END) AS submitted_count,
                   SUM(CASE WHEN result = 'Not Submitted' THEN 1 ELSE 0 END) AS not_submitted_count
            FROM results
            GROUP BY name
        """)
        results = cursor.fetchall()

        # Create a new workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Memberwise Total Report"

        # Add header row
        headers = ["Name", "Submitted Count", "Not Submitted Count"]
        ws.append(headers)

        # Apply bold font to the header row
        for cell in ws[1]:
            cell.font = Font(bold=True)

        # Add data
        for row in results:
            ws.append(row)

        # Apply borders to the entire table
        thin_border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))

        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=3):
            for cell in row:
                cell.border = thin_border

        # Save the workbook
        wb.save(output_path)

        conn.close()
    except Exception as e:
        logging.error(f"Error generating memberwise total report: {e}")
        raise

@app.route('/download_memberwise_total')
def download_memberwise_total():
    try:
        memberwise_total_path = os.path.join(app.config['RESULT_FOLDER'], 'Memberwise_Total_Count.xlsx')
        generate_memberwise_total_report(memberwise_total_path)
        return send_file(memberwise_total_path, as_attachment=True)
    except Exception as e:
        logging.error(f"Error generating memberwise total report: {e}")
        flash('An error occurred while generating the memberwise total report.')
        return redirect('/')



if __name__ == "__main__":
    from waitress import serve
    port = int(os.environ.get('PORT', 8000))
    serve(app, host='0.0.0.0', port=port)


